from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from backend.database.db import get_db
from backend.database.models import User, Lead, Activity
from backend.services.auth_service import require_admin_or_team_leader
import csv
import io
import openpyxl

router = APIRouter(prefix="/api/leads/import", tags=["import"])

# Maps common column name variations to our field names
COLUMN_MAP = {
    "full_name": ["full_name", "name", "full name", "client name", "client", "lead name", "contact"],
    "phone": ["phone", "phone number", "mobile", "mobile number", "tel", "telephone", "contact number"],
    "email": ["email", "email address", "e-mail", "mail"],
    "source": ["source", "lead source", "channel"],
    "budget": ["budget", "budget (aed)", "price", "price range", "aed"],
    "property_type": ["property_type", "property type", "type", "unit type", "property"],
    "preferred_area": ["preferred_area", "preferred area", "area", "location", "community", "preferred location"],
    "notes": ["notes", "note", "remarks", "comment", "comments", "description"],
}

def detect_column(header: str) -> str | None:
    clean = header.strip().lower()
    for field, variants in COLUMN_MAP.items():
        if clean in variants:
            return field
    return None

def parse_rows(headers: list[str], rows: list[list]) -> tuple[list[dict], list[str]]:
    mapping = {}
    for i, h in enumerate(headers):
        field = detect_column(h)
        if field:
            mapping[field] = i

    unrecognized = [h for h in headers if detect_column(h) is None]
    leads = []
    for row in rows:
        if not any(str(c).strip() for c in row):
            continue
        lead = {}
        for field, idx in mapping.items():
            val = str(row[idx]).strip() if idx < len(row) else ""
            if val and val.lower() not in ("none", "null", "n/a", "-", ""):
                lead[field] = val
        if lead.get("full_name") or lead.get("phone"):
            leads.append(lead)

    return leads, unrecognized


@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin_or_team_leader),
):
    content = await file.read()
    filename = file.filename.lower()

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty file")
        headers = all_rows[0]
        rows = all_rows[1:]

    elif filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = [[str(c.value) if c.value is not None else "" for c in row] for row in ws.iter_rows()]
        wb.close()
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty file")
        headers = all_rows[0]
        rows = all_rows[1:]
    else:
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported")

    leads, unrecognized = parse_rows(headers, rows)

    return {
        "headers": headers,
        "column_mapping": {h: detect_column(h) for h in headers},
        "unrecognized_columns": unrecognized,
        "preview": leads[:5],
        "total_rows": len(rows),
        "importable_rows": len(leads),
    }


@router.post("")
async def import_leads(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin_or_team_leader),
    db: Session = Depends(get_db),
):
    content = await file.read()
    filename = file.filename.lower()

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty file")
        headers = all_rows[0]
        rows = all_rows[1:]

    elif filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = [[str(c.value) if c.value is not None else "" for c in row] for row in ws.iter_rows()]
        wb.close()
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty file")
        headers = all_rows[0]
        rows = all_rows[1:]
    else:
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported")

    leads_data, _ = parse_rows(headers, rows)

    created = 0
    skipped = 0
    errors = []

    for i, data in enumerate(leads_data):
        try:
            if not data.get("full_name") and not data.get("phone"):
                skipped += 1
                continue

            lead = Lead(
                full_name=data.get("full_name", "Unknown"),
                phone=data.get("phone", ""),
                email=data.get("email"),
                source=data.get("source", "Other"),
                budget=data.get("budget"),
                property_type=data.get("property_type"),
                preferred_area=data.get("preferred_area"),
                notes=data.get("notes"),
                stage="new_lead",
                assigned_to=current_user.id,
                created_by=current_user.id,
            )
            db.add(lead)
            db.flush()

            activity = Activity(
                lead_id=lead.id,
                user_id=current_user.id,
                type="note",
                content=f"Lead imported via CSV/Excel by {current_user.full_name}",
            )
            db.add(activity)
            created += 1
        except Exception as e:
            skipped += 1
            errors.append(f"Row {i + 2}: {str(e)}")

    db.commit()

    return {
        "ok": True,
        "created": created,
        "skipped": skipped,
        "errors": errors[:10],
    }
